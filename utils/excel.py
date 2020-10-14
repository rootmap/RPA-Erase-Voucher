from typing import NewType
import openpyxl
from openpyxl.formula.translate import Translator


class Excel:
    _Worksheet = NewType('Worksheet', openpyxl)
    _Workbook = NewType('Workbook', openpyxl)

    def _col2num(self, column: str) -> int:
        num = 0
        for c in column:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num

    def open_excel_book(self, location: str) -> _Workbook:
        workbook = openpyxl.load_workbook(location)
        return workbook

    # Copy range of cells as a nested list
    # Takes: start cell, end cell, and sheet you want to copy from.
    def copy_range(self, start_cell: str, end_cell: str, sheet: _Worksheet) -> list:
        start_col = self._col2num(''.join(filter(str.isalpha, start_cell)))
        start_row = int(''.join(filter(str.isnumeric, start_cell)))
        end_col = self._col2num(''.join(filter(str.isalpha, end_cell)))
        end_row = int(''.join(filter(str.isnumeric, end_cell)))
        range_selected = []
        # Loops through selected Rows
        for i in range(start_row, end_row + 1, 1):
            # Appends the row to a RowSelected list
            row_selected = []
            for j in range(start_col, end_col + 1, 1):
                row_selected.append(sheet.cell(row=i, column=j).value)
            # Adds the RowSelected List and nests inside the rangeSelected
            range_selected.append(row_selected)

        return range_selected

    def paste_range(self, start_cell: str, end_cell: str, sheet_receiving: _Worksheet,
                    copied_data: list):
        start_col = self._col2num(''.join(filter(str.isalpha, start_cell)))
        start_row = int(''.join(filter(str.isnumeric, start_cell)))
        end_col = self._col2num(''.join(filter(str.isalpha, end_cell)))
        end_row = int(''.join(filter(str.isnumeric, end_cell)))
        count_row = 0
        for i in range(start_row, end_row + 1, 1):
            count_col = 0
            for j in range(start_col, end_col + 1, 1):
                sheet_receiving.cell(row=i, column=j).value = copied_data[count_row][count_col]
                count_col += 1
            count_row += 1

    def translate_formula(self, sheetname, workbook, source, target):
        # worksheet = workbook[sheetname]  # Add Sheet name
        formula = sheetname[source].value
        sheetname[target] = Translator(formula, origin=source).translate_formula(target)


# xx = Excel()
#
# xx.translate_formula('Sheet1', r"E:\Python Projects\excel\test.xlsx", 'B8', 'C8')
